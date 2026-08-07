#!/usr/bin/env python3
"""Tiny local server for the ledger dashboard.

Serves this directory over HTTP (so the browser can fetch transactions.csv
without file:// permission prompts), accepts POSTed edits back to disk, and
handles statement uploads (CSV / Excel / PDF, including password-protected
Excel and PDF files).

Run (from this directory): .venv/bin/python3 server.py
Then open: http://localhost:8765/spend_analyser.html

Runs with Flask's debug reloader, so editing this file restarts the server
automatically — no manual restart needed after a code change.

Uses the .venv interpreter so openpyxl/msoffcrypto-tool/pdfplumber/pikepdf
(installed there) are importable — running with a bare system python3 will
work for CSV uploads and the dashboard, but Excel/PDF uploads need the venv.
"""
import csv
import io
import json
import os
import re
import uuid
from datetime import datetime

from flask import Flask, jsonify, request, send_from_directory

try:
    import openpyxl
    import msoffcrypto
except ImportError:
    openpyxl = None
    msoffcrypto = None

try:
    import pdfplumber
    import pikepdf
except ImportError:
    pdfplumber = None
    pikepdf = None

PORT = 8765
DIRECTORY = os.path.dirname(os.path.abspath(__file__))
CSV_PATH = os.path.join(DIRECTORY, "transactions.csv")
RULES_PATH = os.path.join(DIRECTORY, "category_rules.json")
INCOMING_DIR = os.path.join(DIRECTORY, "incoming_statements")
MANIFEST_PATH = os.path.join(INCOMING_DIR, "manifest.json")
SUBJECTS_PATH = os.path.join(DIRECTORY, "statement_subjects.json")
os.makedirs(INCOMING_DIR, exist_ok=True)

CSV_FIELDS = ["id", "date", "description", "amount", "account", "category"]
CATEGORIES = [
    "Food & Dining", "Groceries", "Transport", "Shopping", "Bills & Utilities",
    "Rent", "Entertainment", "Health & Fitness", "Subscriptions", "Travel",
    "Insurance", "Loan / EMI", "Investments", "Dividend", "Bank Charges",
    "Salary / Income", "Cashback", "Self Transfer", "Family Transfer",
    "Credit Card Bill", "Settlement", "Transfers", "Others",
]

# Enforced so the account name alone always reveals whether it's a bank/savings
# account or a credit card — the dashboard's account-type filter relies on this
# suffix rather than a separate stored field.
ACCOUNT_LABEL_SUFFIXES = ("Credit Card", "Saving Account Statement")


def validate_account_label(account):
    """Returns an error message string if invalid, None if OK."""
    if not account:
        return "Account label is required."
    if not account.endswith(ACCOUNT_LABEL_SUFFIXES):
        return 'Account label must end with "Credit Card" or "Saving Account Statement".'
    return None


# These appear as a CREDIT on the credit card's own statement (payment received,
# reducing the balance owed) — not a purchase, so they aren't spend or income on the
# card side. Skip them at ingestion rather than let them inflate the "Income" figure.
# Distinct from the "Credit Card Bill" category below, which is the corresponding DEBIT
# on the *bank account* statement (money actually leaving a tracked account to pay the
# bill) — that one is real spend and stays in the ledger.
EXCLUDE_KEYWORDS = ["bppy cc payment", "bbps payment received"]


def load_rules():
    with open(RULES_PATH, encoding="utf-8") as f:
        return json.load(f)


def load_manifest():
    if not os.path.exists(MANIFEST_PATH):
        return {}
    with open(MANIFEST_PATH, encoding="utf-8") as f:
        return json.load(f)


def save_manifest(manifest):
    with open(MANIFEST_PATH, "w", encoding="utf-8") as f:
        json.dump(manifest, f, indent=2)


def load_subjects():
    if not os.path.exists(SUBJECTS_PATH):
        return {}
    with open(SUBJECTS_PATH, encoding="utf-8") as f:
        return json.load(f)


def save_subjects(subjects):
    with open(SUBJECTS_PATH, "w", encoding="utf-8") as f:
        json.dump(subjects, f, indent=2)
        f.write("\n")


def guess_category(desc, rules):
    d = (desc or "").lower()
    for cat, keywords in rules.items():
        if any(kw in d for kw in keywords):
            return cat
    return "Others"


def parse_amount(v):
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    neg = s.startswith("(") and s.endswith(")")
    if neg:
        s = s[1:-1]
    # Strip everything except digits/sign/decimal point/comma — handles currency
    # symbols and OCR/font artifacts (e.g. some HDFC-generated PDFs render ₹ as
    # a literal "C" in extracted text).
    s = re.sub(r"[^\d+\-.,]", "", s).replace(",", "")
    if not s:
        return None
    try:
        n = float(s)
    except ValueError:
        return None
    return -n if neg else n


DATE_FORMATS = ["%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y", "%d.%m.%Y", "%d/%m/%y", "%d-%m-%y"]


def parse_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    s = str(v).strip()
    if not s:
        return None
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(s).date().isoformat()
    except ValueError:
        return None


def rows_to_transactions(rows, account, rules):
    rows = [r for r in rows if any(str(c).strip() for c in r)]
    if not rows:
        return []

    header = [str(c or "").strip().lower() for c in rows[0]]

    def find(*keys):
        for i, c in enumerate(header):
            if any(k in c for k in keys):
                return i
        return -1

    looks_like_header = any(
        find(k) >= 0 for k in ["date", "desc", "narration", "particular", "amount", "debit", "credit", "withdrawal", "deposit"]
    )
    if looks_like_header:
        idx = {
            "date": find("date"),
            "desc": find("desc", "narration", "particular", "detail"),
            "debit": find("debit", "withdrawal"),
            "credit": find("credit", "deposit"),
            "amount": find("amount"),
            "category": find("category"),
        }
        data_rows = rows[1:]
    else:
        idx = {"date": 0, "desc": 1, "amount": 2, "debit": -1, "credit": -1, "category": -1}
        data_rows = rows

    out = []
    for r in data_rows:
        if all(not str(c).strip() for c in r):
            continue

        def cell(i):
            return r[i] if 0 <= i < len(r) else None

        date_val = parse_date(cell(idx["date"]))
        desc = str(cell(idx["desc"]) or "").strip() if idx["desc"] >= 0 else str(cell(1) or "").strip()
        amount = None
        if idx["amount"] >= 0:
            amount = parse_amount(cell(idx["amount"]))
        elif idx["debit"] >= 0 or idx["credit"] >= 0:
            deb = parse_amount(cell(idx["debit"])) if idx["debit"] >= 0 else None
            cred = parse_amount(cell(idx["credit"])) if idx["credit"] >= 0 else None
            if deb is not None and deb > 0:
                amount = -abs(deb)
            elif cred is not None and cred > 0:
                amount = abs(cred)
        if not date_val or amount is None or not desc:
            continue
        if any(kw in desc.lower() for kw in EXCLUDE_KEYWORDS):
            continue
        provided_cat = str(cell(idx["category"]) or "").strip() if idx["category"] >= 0 else ""
        category = provided_cat if provided_cat in CATEGORIES else guess_category(desc, rules)
        out.append({
            "id": uuid.uuid4().hex[:8],
            "date": date_val,
            "description": desc,
            "amount": amount,
            "account": account or "Unlabelled",
            "category": category,
        })
    return out


def read_transactions():
    if not os.path.exists(CSV_PATH):
        return []
    with open(CSV_PATH, encoding="utf-8") as f:
        return list(csv.DictReader(f))


def write_transactions(rows):
    with open(CSV_PATH, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_FIELDS, quoting=csv.QUOTE_ALL)
        writer.writeheader()
        for r in sorted(rows, key=lambda r: r["date"]):
            writer.writerow({k: r.get(k, "") for k in CSV_FIELDS})


def transaction_signature(t):
    """Content-based identity for a transaction, ignoring the id (which is
    randomly generated fresh on every parse, so it can't be used to detect
    that the same statement was uploaded twice)."""
    return (t["date"], t["description"].strip(), round(float(t["amount"]), 2), t["account"].strip())


class NeedsPasswordError(Exception):
    pass


class WrongPasswordError(Exception):
    pass


def read_csv_rows(data):
    text = data.decode("utf-8-sig", errors="replace")
    if not text.strip():
        return []
    try:
        dialect = csv.Sniffer().sniff(text.splitlines()[0], delimiters=",\t;|")
        delim = dialect.delimiter
    except csv.Error:
        delim = ","
    return list(csv.reader(io.StringIO(text), delimiter=delim))


def read_excel_rows(data, password):
    if openpyxl is None:
        raise RuntimeError("openpyxl not installed — run: source .venv/bin/activate && pip install openpyxl msoffcrypto-tool")
    buf = io.BytesIO(data)
    try:
        wb = openpyxl.load_workbook(buf, read_only=True, data_only=True)
    except Exception:
        if msoffcrypto is None:
            raise RuntimeError("msoffcrypto-tool not installed")
        buf.seek(0)
        office_file = msoffcrypto.OfficeFile(buf)
        if not password:
            raise NeedsPasswordError()
        try:
            office_file.load_key(password=password)
            decrypted = io.BytesIO()
            office_file.decrypt(decrypted)
        except Exception:
            raise WrongPasswordError()
        decrypted.seek(0)
        wb = openpyxl.load_workbook(decrypted, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    return [list(row) for row in ws.iter_rows(values_only=True)]


DATE_LINE_RE = re.compile(r"^(\d{1,2}[/\-.]\d{1,2}[/\-.]\d{2,4})(?:\s*[|,]?\s*\d{1,2}:\d{2})?\s+(.*)$")
TRAILING_AMOUNT_RE = re.compile(r"([+\-]?\s*[₹C$]?\s*[\d,]+\.\d{1,2})\s*[a-zA-Z]{0,2}\s*$")


def parse_statement_lines(lines):
    """Best-effort: pull (date, description, amount) triples out of raw PDF
    text lines. Works for simple 'date  description  amount' layouts. Complex
    multi-column / multi-cardholder statements (e.g. credit card statements
    with a separate rewards-points column) are NOT reliably handled this way —
    for those, upload the PDF to Claude Code directly in chat instead, where it
    can be read and cross-checked against the statement's own totals."""
    rows = []
    for line in lines:
        line = line.strip()
        if not line:
            continue
        m = DATE_LINE_RE.match(line)
        if not m:
            continue
        date_str, rest = m.group(1), m.group(2)
        amt_m = TRAILING_AMOUNT_RE.search(rest)
        if not amt_m:
            continue
        desc = rest[:amt_m.start()].strip()
        if not desc:
            continue
        raw_amt = amt_m.group(1).strip()
        # These statements mark credits with a leading "+" and leave debits
        # unsigned — so an unsigned amount here means spend, not the usual
        # "no sign = positive" convention. Make the sign explicit before this
        # flows into the shared parse_amount().
        is_credit = raw_amt.startswith("+")
        numeric = re.sub(r"[^\d.,]", "", raw_amt)
        signed = numeric if is_credit else "-" + numeric
        rows.append([date_str, desc, signed])
    return rows


BANK_DATE_LINE_RE = re.compile(r"^(\d{1,2}/\d{1,2}/\d{4})\s+(.*)$")
BANK_TRAILING_NUMS_RE = re.compile(r"([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s*$")

# Repeated page header/footer boilerplate and end-of-table sections — a
# continuation line matching any of these means we've left the transaction
# table (page break, summary, FD details, etc.), so stop accumulating rather
# than let it bleed into the next/previous transaction's narration.
BANK_STOP_MARKERS = [
    "page ", "customer id", "account branch", "statement from", "joint holders",
    "nomination", "expected amb", "opening balance", "txn date", "narration",
    "summary", "debit count", "total withdrawal balance", "total sweep in fd",
    "fd details", "fd number", "current principal", "available withdrawable",
    "details of td interest", "interest paid/", "disclaimer", "end of statement",
]


def _is_bank_stop_marker(line):
    d = line.lower()
    return any(marker in d for marker in BANK_STOP_MARKERS)


def parse_bank_statement_lines(lines):
    """Bank/savings account statements: 3 numeric columns (withdrawal, deposit,
    closing balance) per transaction — unlike credit card statements' single
    amount column. Narration text often wraps onto several following lines
    with no numbers on them; only the first line carries the amounts, but the
    continuation lines often carry the actual classifying detail (e.g. a
    salary credit's "Toast Tab Salary Jun-26" text appears on a continuation
    line, not the first line) — so those get appended to the narration rather
    than dropped. Continuation lines that look like page header/footer
    boilerplate stop the accumulation instead of bleeding into the record.
    Returns data rows only (no header) — the caller decides which parser won."""
    rows = []
    current = None

    def flush():
        if current:
            rows.append([current["date"], " ".join(current["parts"]), current["withdrawal"], current["deposit"]])

    for line in lines:
        line = line.strip()
        if not line:
            continue
        m = BANK_DATE_LINE_RE.match(line)
        nums_m = BANK_TRAILING_NUMS_RE.search(m.group(2)) if m else None
        if m and nums_m:
            flush()
            desc = m.group(2)[:nums_m.start()].strip()
            withdrawal, deposit, _closing = nums_m.groups()
            current = {"date": m.group(1), "parts": [desc] if desc else [], "withdrawal": withdrawal, "deposit": deposit}
        elif current and _is_bank_stop_marker(line):
            flush()
            current = None
        elif current:
            current["parts"].append(m.group(2) if m else line)
    flush()
    return rows


AXIS_DATE_LINE_RE = re.compile(r"^(\d{1,2}/\d{1,2}/\d{4})\s+(.*)$")
AXIS_TRAILING_AMOUNT_RE = re.compile(r"([\d,]+\.\d{2})\s+(Dr|Cr)\s*$", re.IGNORECASE)


def parse_axis_statement_lines(lines):
    """Statements (e.g. Axis Bank) that suffix each amount with Dr/Cr instead
    of a leading +/blank sign convention. Single line per transaction, no
    wrapping — simpler than the bank-statement format. Returns data rows only
    (no header) — the caller decides which parser won."""
    rows = []
    for line in lines:
        line = line.strip()
        if not line:
            continue
        m = AXIS_DATE_LINE_RE.match(line)
        if not m:
            continue
        date_str, rest = m.group(1), m.group(2)
        amt_m = AXIS_TRAILING_AMOUNT_RE.search(rest)
        if not amt_m:
            continue
        desc = rest[:amt_m.start()].strip()
        if not desc:
            continue
        amount_str, sign = amt_m.groups()
        numeric = amount_str.replace(",", "")
        signed = numeric if sign.lower() == "cr" else "-" + numeric
        rows.append([date_str, desc, signed])
    return rows


AXIS_SAVINGS_DATE_LINE_RE = re.compile(r"^(\d{2}-\d{2}-\d{4})\s+(.*)$")
AXIS_SAVINGS_TRAILING_TWO_RE = re.compile(r"([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s*$")

AXIS_SAVINGS_STOP_MARKERS = [
    "important message", "important notice", "disclaimer", "legends used",
    "rejection code", "page ", "statement for account", "date transaction",
    "deposit insurance", "in compliance with regulatory", "total ",
]


def _is_axis_savings_stop_marker(line):
    d = line.lower()
    return any(marker in d for marker in AXIS_SAVINGS_STOP_MARKERS)


def parse_axis_savings_statement_lines(lines):
    """Axis Bank savings/salary account statements: each transaction line
    carries only two trailing numbers (amount, running balance) instead of a
    fixed withdrawal/deposit column pair — whichever of withdrawal/deposit
    doesn't apply is left blank rather than printed as 0.00. So there's no
    column to read the direction from; instead track the running balance and
    infer debit vs credit from which way it moved."""
    rows = []
    current = None  # index into rows of the transaction being accumulated
    prev_balance = 0.0

    for line in lines:
        line = line.strip()
        if not line:
            continue
        low = line.lower()
        if low.startswith("opening balance"):
            m = re.search(r"([\d,]+\.\d{2})", line)
            if m:
                prev_balance = float(m.group(1).replace(",", ""))
            current = None
            continue
        if low.startswith("closing balance"):
            current = None
            continue
        m = AXIS_SAVINGS_DATE_LINE_RE.match(line)
        if m:
            current = None
            rest = m.group(2)
            nums_m = AXIS_SAVINGS_TRAILING_TWO_RE.search(rest)
            if not nums_m:
                continue
            desc = rest[:nums_m.start()].strip()
            amount = float(nums_m.group(1).replace(",", ""))
            new_balance = float(nums_m.group(2).replace(",", ""))
            diff = round(new_balance - prev_balance, 2)
            signed = amount if abs(diff - amount) < 0.01 else -amount
            prev_balance = new_balance
            rows.append([m.group(1), desc, f"{signed:.2f}"])
            current = len(rows) - 1
            continue
        if current is not None and _is_axis_savings_stop_marker(line):
            current = None
            continue
        if current is not None:
            rows[current][1] = (rows[current][1] + " " + line).strip()

    return rows


ICICI_DATE_LINE_RE = re.compile(r"^(\d{2}-\d{2}-\d{4})(.*)$")
ICICI_TRAILING_TWO_RE = re.compile(r"([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s*$")
ICICI_BF_RE = re.compile(r"^\d{2}-\d{2}-\d{4}\s+B/F\s+([\d,]+\.\d{2})\s*$")

ICICI_STOP_MARKERS = [
    "statement of transactions", "date mode particulars", "page ", "total:",
    "sincerely", "legends for transactions", "this is a system generated",
    "you can now download", "customers with pradhan", "card blocking",
    "account blocking", "w.e.f.", "icici bank will follow", "update nominee",
]


def _is_icici_stop_marker(line):
    d = line.lower()
    return any(marker in d for marker in ICICI_STOP_MARKERS)


def parse_icici_statement_lines(lines):
    """ICICI Bank savings account statements: unlike the other 3-line-column or
    2-trailing-number formats, the narration here isn't a clean run of
    continuation lines after the amount — the payee name and UPI reference
    text wrap both BEFORE and AFTER the date/amount line, e.g. a payee name
    line, then a UPI-detail line, then the date+amount+balance line, then more
    reference-code lines, then the NEXT payee name. There's no reliable way to
    tell exactly where one transaction's trailing reference code ends and the
    next one's leading payee name begins from the text alone, so this
    attributes all buffered non-date lines to the transaction whose date/
    amount line follows them (i.e. treats them as leading narration for the
    next row) — this means a little bit of the previous row's leftover
    reference text ends up prefixed onto the next row's description, but the
    amounts/balances (validated by reconciling against the statement's own
    per-page "Total:" subtotals) are unaffected either way."""
    rows = []
    buf = []
    prev_balance = None

    for line in lines:
        line = line.strip()
        if not line:
            continue
        bf = ICICI_BF_RE.match(line)
        if bf:
            prev_balance = float(bf.group(1).replace(",", ""))
            buf = []
            continue
        if _is_icici_stop_marker(line):
            buf = []
            continue
        m = ICICI_DATE_LINE_RE.match(line)
        if m:
            rest = m.group(2)
            nums_m = ICICI_TRAILING_TWO_RE.search(rest)
            if not nums_m:
                buf.append(line)
                continue
            inline = rest[:nums_m.start()].strip()
            amount = float(nums_m.group(1).replace(",", ""))
            new_balance = float(nums_m.group(2).replace(",", ""))
            desc = " ".join(buf + ([inline] if inline else [])).strip()
            if prev_balance is not None:
                diff = round(new_balance - prev_balance, 2)
                signed = amount if abs(diff - amount) < 0.01 else -amount
            else:
                signed = amount
            prev_balance = new_balance
            rows.append([m.group(1), desc, f"{signed:.2f}"])
            buf = []
            continue
        buf.append(line)

    return rows


ICICI_CARD_DATE_LINE_RE = re.compile(r"^(\d{2}/\d{2}/\d{4})\s+(.*)$")
ICICI_CARD_TRAILING_AMOUNT_RE = re.compile(r"([\d,]+\.\d{2})\s*(CR)?\s*$", re.IGNORECASE)
ICICI_CARD_SERNO_RE = re.compile(r"^\d+\s+(.*)$")
ICICI_CARD_TRAILING_POINTS_RE = re.compile(r"\s+\d+\s*$")


def parse_icici_card_statement_lines(lines):
    """ICICI Bank credit card statements (e.g. Sapphiro): single line per
    transaction, `date serno description reward-points amount [CR]` — a
    trailing CR marks a payment/credit, its absence marks a purchase/debit,
    same sign convention as parse_axis_statement_lines()."""
    rows = []
    for line in lines:
        line = line.strip()
        if not line:
            continue
        m = ICICI_CARD_DATE_LINE_RE.match(line)
        if not m:
            continue
        date_str, rest = m.group(1), m.group(2)
        amt_m = ICICI_CARD_TRAILING_AMOUNT_RE.search(rest)
        if not amt_m:
            continue
        amount_str, cr = amt_m.groups()
        desc_and_points = rest[:amt_m.start()].strip()
        pts_m = ICICI_CARD_TRAILING_POINTS_RE.search(desc_and_points)
        if pts_m:
            desc_and_points = desc_and_points[:pts_m.start()]
        serno_m = ICICI_CARD_SERNO_RE.match(desc_and_points)
        desc = serno_m.group(1).strip() if serno_m else desc_and_points.strip()
        if not desc:
            continue
        numeric = amount_str.replace(",", "")
        signed = numeric if cr else "-" + numeric
        rows.append([date_str, desc, signed])
    return rows


def extract_pdf_lines(data, password):
    """Decrypt + extract raw text lines from a PDF, with no format-specific
    parsing — used both as the first step of read_pdf_rows() and standalone
    by the format-check diagnostic route, so an unrecognized layout can be
    inspected without duplicating the pikepdf/pdfplumber/cid-stripping
    logic."""
    if pdfplumber is None or pikepdf is None:
        raise RuntimeError("pdfplumber/pikepdf not installed — run: source .venv/bin/activate && pip install pdfplumber pikepdf")
    try:
        pdf = pikepdf.open(io.BytesIO(data), password=password or "")
    except pikepdf.PasswordError:
        if not password:
            raise NeedsPasswordError()
        raise WrongPasswordError()
    buf = io.BytesIO()
    pdf.save(buf)
    buf.seek(0)

    lines = []
    with pdfplumber.open(buf) as doc:
        for page in doc.pages:
            text = page.extract_text() or ""
            # Some statements' fonts map a tab glyph through a cmap that
            # pdfplumber can't resolve to whitespace, so it falls back to
            # printing the raw "(cid:9)" placeholder inline — e.g. this shows
            # up as "Transaction(cid:9)Details" instead of "Transaction
            # Details". Same class of bug as the "C"-for-₹ OCR artifact:
            # silently corrupts every line unless stripped first.
            text = text.replace("(cid:9)", " ")
            lines.extend(text.split("\n"))
    return lines


def read_pdf_rows(data, password):
    lines = extract_pdf_lines(data, password)

    # Different banks format transaction lines very differently (single
    # amount column with a leading +/blank sign, a 3-column withdrawal/
    # deposit/balance layout, a trailing Dr/Cr suffix, or two trailing
    # numbers — amount + running balance — with direction inferred from the
    # balance). Loose keyword sniffing (e.g. "withdrawals"/"deposits"
    # anywhere in the text) is unreliable — dense disclaimer boilerplate can
    # contain those words incidentally, and picking the parser with the most
    # matched rows is also unreliable (a wrong-format parser can still
    # spuriously match many lines using the wrong number as the amount).
    # Instead, check for each format's actual column-header phrase, which
    # only appears in the real table header, not in unrelated prose.
    full_text_lower = "\n".join(lines).lower()
    if "date mode particulars" in full_text_lower:
        return [["date", "description", "amount"]] + parse_icici_statement_lines(lines)
    if "date serno" in full_text_lower:
        return [["date", "description", "amount"]] + parse_icici_card_statement_lines(lines)
    if "transaction details" in full_text_lower and "withdrawal deposits balance" in full_text_lower:
        return [["date", "description", "amount"]] + parse_axis_savings_statement_lines(lines)
    if "narration withdrawals deposits" in full_text_lower:
        return [["date", "narration", "withdrawals", "deposits"]] + parse_bank_statement_lines(lines)
    if "merchant category" in full_text_lower:
        return [["date", "description", "amount"]] + parse_axis_statement_lines(lines)

    return [["date", "description", "amount"]] + parse_statement_lines(lines)


def parse_statement(data, filename, account, password=None):
    """Parse+categorize a statement's raw bytes into transaction dicts,
    WITHOUT touching transactions.csv — used both as the first step of
    ingest_statement() and standalone for the pending-statements preview.
    Returns {"transactions": [...]}, {"needs_password": True}, or
    {"error": "..."}."""
    try:
        name = filename.lower()
        if name.endswith(".csv"):
            rows = read_csv_rows(data)
        elif name.endswith(".xlsx") or name.endswith(".xls"):
            rows = read_excel_rows(data, password)
        elif name.endswith(".pdf"):
            rows = read_pdf_rows(data, password)
        else:
            return {"error": "Unsupported file type — only .csv, .xlsx, .xls, .pdf are supported."}

        rules = load_rules()
        txs = rows_to_transactions(rows, account, rules)
        if not txs:
            msg = "No valid transaction rows found in this file."
            if name.endswith(".pdf"):
                msg += " PDF parsing is best-effort and doesn't handle complex/multi-column statements — paste this PDF to Claude Code in chat instead."
            return {"error": msg}
        return {"transactions": txs}

    except NeedsPasswordError:
        return {"needs_password": True}
    except WrongPasswordError:
        return {"error": "Incorrect password."}


def ingest_statement(data, filename, account, password=None):
    """Parse+categorize+dedup+persist a statement's raw bytes into
    transactions.csv. Shared by the browser-upload route and the
    pending-statements (fetched-from-email) route so both go through
    identical logic. Returns {"added": N, "duplicates_skipped": N?} on
    success, {"needs_password": True}, or {"error": "..."}."""
    result = parse_statement(data, filename, account, password)
    if "error" in result or "needs_password" in result:
        return result

    new_txs = result["transactions"]
    existing = read_transactions()
    existing_sigs = {transaction_signature(r) for r in existing}
    unique_new = [t for t in new_txs if transaction_signature(t) not in existing_sigs]
    duplicate_count = len(new_txs) - len(unique_new)

    if not unique_new:
        return {"error": f"All {len(new_txs)} transaction(s) in this file are already in transactions.csv — looks like this statement was uploaded before."}

    write_transactions(existing + unique_new)
    out = {"added": len(unique_new)}
    if duplicate_count:
        out["duplicates_skipped"] = duplicate_count
    return out


app = Flask(__name__, static_folder=None)


@app.after_request
def _quiet_and_uncached(resp):
    resp.headers["Cache-Control"] = "no-store"
    return resp


@app.route("/")
@app.route("/<path:filename>")
def serve_static(filename="spend_analyser.html"):
    return send_from_directory(DIRECTORY, filename)


@app.route("/accounts")
def accounts():
    return jsonify(sorted({r["account"] for r in read_transactions() if r.get("account")}))


@app.route("/save", methods=["POST"])
def save():
    body = request.get_data(as_text=True)
    with open(CSV_PATH, "w", encoding="utf-8") as f:
        f.write(body)
    return "ok"


@app.route("/upload", methods=["POST"])
def upload():
    try:
        file_item = request.files.get("file")
        if file_item is None:
            return jsonify({"error": "No file provided."}), 400
        filename = file_item.filename or ""
        data = file_item.read()
        password = request.form.get("password") or None
        account = (request.form.get("account") or "").strip()
        account_error = validate_account_label(account)
        if account_error:
            return jsonify({"error": account_error}), 400

        result = ingest_statement(data, filename, account, password)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    status = 200 if ("added" in result or "needs_password" in result) else 400
    return jsonify(result), status


@app.route("/pending")
def pending():
    manifest = load_manifest()
    items = []
    for fn in sorted(os.listdir(INCOMING_DIR)):
        if fn == "manifest.json" or fn.startswith("."):
            continue
        path = os.path.join(INCOMING_DIR, fn)
        if not os.path.isfile(path):
            continue
        meta = manifest.get(fn, {})
        st = os.stat(path)
        items.append({
            "filename": fn,
            "suggested_account": meta.get("account"),
            "subject": meta.get("subject"),
            "received": meta.get("received"),
            "size": st.st_size,
            "mtime": st.st_mtime,
        })
    return jsonify(items)


@app.route("/pending/preview", methods=["POST"])
def pending_preview():
    body = request.get_json(silent=True) or {}
    filename = (body.get("filename") or "").strip()
    account = (body.get("account") or "").strip() or "Preview"
    password = body.get("password") or None
    if not filename:
        return jsonify({"error": "No filename provided."}), 400

    safe_name = os.path.basename(filename)
    path = os.path.join(INCOMING_DIR, safe_name)
    if safe_name != filename or not os.path.isfile(path):
        return jsonify({"error": "File not found in pending folder."}), 404

    try:
        with open(path, "rb") as f:
            data = f.read()
        result = parse_statement(data, safe_name, account, password)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    status = 200 if "transactions" in result or "needs_password" in result else 400
    return jsonify(result), status


@app.route("/pending/upload", methods=["POST"])
def pending_upload():
    body = request.get_json(silent=True) or {}
    filename = (body.get("filename") or "").strip()
    account = (body.get("account") or "").strip()
    password = body.get("password") or None
    if not filename:
        return jsonify({"error": "No filename provided."}), 400
    account_error = validate_account_label(account)
    if account_error:
        return jsonify({"error": account_error}), 400

    safe_name = os.path.basename(filename)
    path = os.path.join(INCOMING_DIR, safe_name)
    if safe_name != filename or not os.path.isfile(path):
        return jsonify({"error": "File not found in pending folder."}), 404

    try:
        with open(path, "rb") as f:
            data = f.read()
        result = ingest_statement(data, safe_name, account, password)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    if "added" in result:
        os.remove(path)
        manifest = load_manifest()
        manifest.pop(safe_name, None)
        save_manifest(manifest)

    status = 200 if ("added" in result or "needs_password" in result) else 400
    return jsonify(result), status


@app.route("/pending/delete", methods=["POST"])
def pending_delete():
    body = request.get_json(silent=True) or {}
    filename = (body.get("filename") or "").strip()
    if not filename:
        return jsonify({"error": "No filename provided."}), 400

    safe_name = os.path.basename(filename)
    path = os.path.join(INCOMING_DIR, safe_name)
    if safe_name != filename or not os.path.isfile(path):
        return jsonify({"error": "File not found in pending folder."}), 404

    os.remove(path)
    manifest = load_manifest()
    manifest.pop(safe_name, None)
    save_manifest(manifest)
    return jsonify({"deleted": safe_name})


@app.route("/pending/fetch-now", methods=["POST"])
def pending_fetch_now():
    try:
        import fetch_statements
    except Exception as e:
        return jsonify({"error": f"Could not load fetch_statements.py: {e}"}), 500
    try:
        summary = fetch_statements.fetch_new_statements()
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    if summary.get("errors"):
        return jsonify({"error": "; ".join(summary["errors"])}), 400
    return jsonify(summary)


@app.route("/format-check", methods=["POST"])
def format_check():
    file_item = request.files.get("file")
    if file_item is None:
        return jsonify({"error": "No file provided."}), 400
    filename = file_item.filename or ""
    data = file_item.read()
    password = request.form.get("password") or None
    account = (request.form.get("account") or "").strip() or "Format check"

    try:
        result = parse_statement(data, filename, account, password)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    if "transactions" in result:
        return jsonify({"transactions": result["transactions"]})
    if "needs_password" in result:
        return jsonify(result)

    # A PDF error here could mean two different things: no known layout
    # matched at all (worth showing raw text to build a new parser), or a
    # layout matched fine and extracted rows, but every row was filtered out
    # downstream (EXCLUDE_KEYWORDS self-payments, or a row missing a date/
    # amount/description) — that second case is not a new-format problem, so
    # check whether read_pdf_rows actually produced rows before treating this
    # as unrecognized.
    if filename.lower().endswith(".pdf") and "Incorrect password" not in result.get("error", ""):
        try:
            rows = read_pdf_rows(data, password)
        except Exception:
            rows = []
        if len(rows) > 1:
            result["error"] = (
                f"A known statement layout matched and extracted {len(rows) - 1} row(s), "
                "but none became usable transactions — likely all excluded as self-payments "
                "(EXCLUDE_KEYWORDS) or missing a date/amount/description. This isn't a "
                "new-format problem, so no subject mapping is needed for this file."
            )
        else:
            try:
                lines = extract_pdf_lines(data, password)
                result["raw_excerpt"] = "\n".join(lines[:400])
            except Exception:
                pass

    return jsonify(result), 400


@app.route("/statement-subjects")
def statement_subjects():
    return jsonify(load_subjects())


@app.route("/statement-subjects/save", methods=["POST"])
def statement_subjects_save():
    body = request.get_json(silent=True) or {}
    phrase = (body.get("phrase") or "").strip()
    account = (body.get("account") or "").strip()
    if not phrase:
        return jsonify({"error": "A subject phrase is required."}), 400
    account_error = validate_account_label(account)
    if account_error:
        return jsonify({"error": account_error}), 400

    subjects = load_subjects()
    subjects[phrase] = account
    save_subjects(subjects)
    return jsonify(subjects)


@app.route("/statement-subjects/delete", methods=["POST"])
def statement_subjects_delete():
    body = request.get_json(silent=True) or {}
    phrase = (body.get("phrase") or "").strip()
    if not phrase:
        return jsonify({"error": "No phrase provided."}), 400

    subjects = load_subjects()
    subjects.pop(phrase, None)
    save_subjects(subjects)
    return jsonify(subjects)


if __name__ == "__main__":
    # LAN access (0.0.0.0) temporarily turned back off — switch host back to
    # "0.0.0.0" to let other devices on the same wifi reach this again.
    print(f"Serving {DIRECTORY} at http://localhost:{PORT}/spend_analyser.html")
    app.run(host="127.0.0.1", port=PORT, debug=True, use_reloader=True)
